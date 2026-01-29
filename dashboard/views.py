from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.conf import settings
from django.utils import timezone
from datetime import timedelta, datetime, time as dt_time
from collections import defaultdict

from account.models import User, AuditLog
from .models import Shift

from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count
from django.contrib import messages
from .forms import (ShiftForm, ShiftFilterForm, BulkShiftActionForm, 
                    UserManagementForm, UserCreateForm, UserFilterForm, 
                    PasswordResetFormAdmin, ShiftExcelUploadForm)
import json

from django.contrib.auth.hashers import make_password

import openpyxl
from openpyxl import Workbook

# Import appointment models
from appointments.models import Appointment

def get_user_shifts(user, days=7):
    """Get upcoming shifts for a user"""
    today = timezone.now().date()
    end_date = today + timedelta(days=days)
    return Shift.objects.filter(
        user=user,
        date__range=[today, end_date]
    ).order_by('date', 'start_time')


def get_today_shifts_by_role(role):
    """Get today's shifts for all users of a specific role"""
    today = timezone.now().date()
    return Shift.objects.filter(
        user__role=role,
        date=today
    ).select_related('user').order_by('start_time')


def get_user_appointments(user, days=7):
    """Get upcoming appointments for a practitioner"""
    today = timezone.now().date()
    end_date = today + timedelta(days=days)
    return Appointment.objects.filter(
        practitioner=user,
        appointment_date__range=[today, end_date],
        status__in=['booked', 'arrived']
    ).select_related('patient', 'service_type').order_by('appointment_date', 'start_time')


def get_today_appointments(user):
    """Get today's appointments for a practitioner"""
    today = timezone.now().date()
    return Appointment.objects.filter(
        practitioner=user,
        appointment_date=today
    ).select_related('patient', 'service_type').order_by('start_time')


@login_required
def dashboard_home(request):
    """Route user to appropriate dashboard based on role"""
    dashboard_name = settings.ROLE_DASHBOARD_MAP.get(request.user.role, 'dashboard:patient')
    return redirect(dashboard_name)


@login_required
def admin_dashboard(request):
    """系統管理員儀表板"""
    if request.user.role != 'admin':
        raise PermissionDenied("您沒有權限訪問此頁面")
    
    today = timezone.now().date()
    
    # Get all shifts for the next 7 days
    end_date = today + timedelta(days=6)
    all_upcoming_shifts = Shift.objects.filter(
        date__range=[today, end_date]
    ).select_related('user').order_by('date', 'start_time', 'user__role')
    
    # Get today's shifts by role
    today_doctors = Shift.objects.filter(
        user__role='doctor',
        date=today
    ).select_related('user').order_by('start_time')
    
    today_therapists = Shift.objects.filter(
        user__role='therapist',
        date=today
    ).select_related('user').order_by('start_time')
    
    today_nurses = Shift.objects.filter(
        user__role='nurse',
        date=today
    ).select_related('user').order_by('start_time')
    
    # Get all staff members
    all_doctors = User.objects.filter(role='doctor', is_active=True)
    all_therapists = User.objects.filter(role='therapist', is_active=True)
    all_nurses = User.objects.filter(role='nurse', is_active=True)
    all_case_managers = User.objects.filter(role='case_manager', is_active=True)
    all_caregivers = User.objects.filter(role='caregiver', is_active=True)
    
    # Organize shifts by date for weekly view
    shifts_by_date = defaultdict(lambda: defaultdict(list))
    
    for shift in all_upcoming_shifts:
        shifts_by_date[shift.date][shift.user.role].append(shift)
    
    # Convert to sorted list of tuples for template
    weekly_schedule = []
    for i in range(7):
        date = today + timedelta(days=i)
        weekly_schedule.append({
            'date': date,
            'is_today': date == today,
            'doctors': shifts_by_date[date].get('doctor', []),
            'therapists': shifts_by_date[date].get('therapist', []),
            'nurses': shifts_by_date[date].get('nurse', []),
        })
    
    # Appointment statistics
    today_appointments = Appointment.objects.filter(appointment_date=today)
    total_appointments_today = today_appointments.count()
    arrived_appointments = today_appointments.filter(status='arrived').count()
    completed_appointments = today_appointments.filter(status='fulfilled').count()
    booked_appointments = today_appointments.filter(status='booked').count()
    
    context = {
        'title': '系統管理儀表板',
        'total_users': User.objects.count(),
        'total_patients': User.objects.filter(role='patient').count(),
        'total_doctors': all_doctors.count(),
        'total_therapists': all_therapists.count(),
        'total_nurses': all_nurses.count(),
        'total_case_managers': all_case_managers.count(),
        'total_caregivers': all_caregivers.count(),
        'recent_users': User.objects.all().order_by('-created_at')[:5],
        'recent_logs': AuditLog.objects.select_related('user').all()[:10],
        
        # Shift information
        'my_shift': Shift.objects.filter(user=request.user, date=today).first(),
        'today_doctors': today_doctors,
        'today_therapists': today_therapists,
        'today_nurses': today_nurses,
        
        # Weekly schedule
        'weekly_schedule': weekly_schedule,
        'all_upcoming_shifts': all_upcoming_shifts,
        
        # Staff lists
        'all_doctors': all_doctors,
        'all_therapists': all_therapists,
        'all_nurses': all_nurses,
        
        # Appointment statistics
        'total_appointments_today': total_appointments_today,
        'arrived_appointments': arrived_appointments,
        'completed_appointments': completed_appointments,
        'booked_appointments': booked_appointments,
    }
    return render(request, 'dashboard/admin_dashboard.html', context)


@login_required
def doctor_dashboard(request):
    """醫師儀表板"""
    if request.user.role != 'doctor':
        raise PermissionDenied("您沒有權限訪問此頁面")
    
    today = timezone.now().date()
    
    today_appts = get_today_appointments(request.user)

    context = {
        'title': '醫師儀表板',
        'doctor': request.user,
        'my_patients_count': User.objects.filter(role='patient').count(),
        
        # Shift information
        'my_shifts': get_user_shifts(request.user, days=7),
        'today_shift': Shift.objects.filter(
            user=request.user,
            date=today
        ).first(),
        
        # Appointment information
        'today_appointments': today_appts,
        'upcoming_appointments': get_user_appointments(request.user, days=7),
        
        # Appointment counts by status
        'booked_count': today_appts.filter(status='booked').count(),
        'arrived_count': today_appts.filter(status='arrived').count(),
        'fulfilled_count': today_appts.filter(status='fulfilled').count(),
    }
    return render(request, 'dashboard/doctor_dashboard.html', context)


@login_required
def therapist_dashboard(request):
    """治療師儀表板"""
    if request.user.role != 'therapist':
        raise PermissionDenied("您沒有權限訪問此頁面")
    
    today = timezone.now().date()
    
    today_appts = get_today_appointments(request.user)

    # Count robot therapy vs regular therapy
    robot_therapy_count = 0
    regular_therapy_count = 0
    for appt in today_appts:
        if appt.service_type.requires_robot:
            robot_therapy_count += 1
        else:
            regular_therapy_count += 1

    context = {
        'title': '治療師儀表板',
        'therapist': request.user,
        'my_patients_count': User.objects.filter(role='patient').count(),
        
        # Shift information
        'my_shifts': get_user_shifts(request.user, days=7),
        'today_shift': Shift.objects.filter(
            user=request.user,
            date=today
        ).first(),
        
        # Appointment information
        'today_appointments': today_appts,
        'upcoming_appointments': get_user_appointments(request.user, days=7),
        
        # Therapy counts
        'robot_therapy_count': robot_therapy_count,
        'regular_therapy_count': regular_therapy_count,
    }
    return render(request, 'dashboard/therapist_dashboard.html', context)


@login_required
def nurse_dashboard(request):
    """護理師儀表板"""
    if request.user.role != 'nurse':
        raise PermissionDenied("您沒有權限訪問此頁面")
    
    today = timezone.now().date()
    
    # Get all today's appointments for check-in management
    all_today_appointments = Appointment.objects.filter(
        appointment_date=today
    ).select_related('patient', 'practitioner', 'service_type').order_by('start_time')
    
    context = {
        'title': '護理師儀表板',
        'nurse': request.user,
        'patients_count': User.objects.filter(role='patient').count(),
        
        # Shift information
        'my_shifts': get_user_shifts(request.user, days=7),
        'today_shift': Shift.objects.filter(
            user=request.user,
            date=today
        ).first(),
        
        # All appointments for check-in
        'all_today_appointments': all_today_appointments,
        'booked_count': all_today_appointments.filter(status='booked').count(),
        'arrived_count': all_today_appointments.filter(status='arrived').count(),
    }
    return render(request, 'dashboard/nurse_dashboard.html', context)


@login_required
def case_manager_dashboard(request):
    """個管師儀表板"""
    if request.user.role != 'case_manager':
        raise PermissionDenied("您沒有權限訪問此頁面")
    
    today = timezone.now().date()
    
    # Get all patients for case management
    all_patients = User.objects.filter(role='patient', is_active=True)
    
    # Get today's appointments for all patients
    today_appointments = Appointment.objects.filter(
        appointment_date=today
    ).select_related('patient', 'practitioner', 'service_type').order_by('start_time')
    
    context = {
        'title': '個管師儀表板',
        'case_manager': request.user,
        'total_patients': all_patients.count(),
        
        # Shift information
        'my_shifts': get_user_shifts(request.user, days=7),
        'today_shift': Shift.objects.filter(
            user=request.user,
            date=today
        ).first(),
        
        # Patient management
        'active_patients': all_patients,
        'today_appointments': today_appointments,
        'total_appointments_today': today_appointments.count(),
    }
    return render(request, 'dashboard/case_manager_dashboard.html', context)


@login_required
def caregiver_dashboard(request):
    """照服員儀表板"""
    if request.user.role != 'caregiver':
        raise PermissionDenied("您沒有權限訪問此頁面")
    
    today = timezone.now().date()
    
    # Get assigned patients (you may need to add assignment logic later)
    assigned_patients = User.objects.filter(role='patient', is_active=True)
    
    context = {
        'title': '照服員儀表板',
        'caregiver': request.user,
        'assigned_patients_count': assigned_patients.count(),
        
        # Shift information
        'my_shifts': get_user_shifts(request.user, days=7),
        'today_shift': Shift.objects.filter(
            user=request.user,
            date=today
        ).first(),
        
        # Patient care
        'assigned_patients': assigned_patients[:10],  # Show first 10
    }
    return render(request, 'dashboard/caregiver_dashboard.html', context)


@login_required
def patient_dashboard(request):
    """病患儀表板"""
    if request.user.role != 'patient':
        raise PermissionDenied("您沒有權限訪問此頁面")
    
    context = {
        'title': '我的健康儀表板',
        'patient': request.user,
    }
    return render(request, 'dashboard/patient_dashboard.html', context)


@login_required
def researcher_dashboard(request):
    """研究員儀表板"""
    if request.user.role != 'researcher':
        raise PermissionDenied("您沒有權限訪問此頁面")
    
    context = {
        'title': '研究員儀表板',
        'researcher': request.user,
        'total_patients': User.objects.filter(role='patient').count(),
    }
    return render(request, 'dashboard/researcher_dashboard.html', context)


@login_required
def shift_management(request):
    """Shift management page for admin"""
    if request.user.role != 'admin':
        raise PermissionDenied("您沒有權限訪問此頁面")
    
    # Get filter parameters
    filter_form = ShiftFilterForm(request.GET or None)
    
    # Base queryset
    shifts = Shift.objects.select_related('user').all()
    
    # Apply filters
    if filter_form.is_valid():
        role = filter_form.cleaned_data.get('role')
        status = filter_form.cleaned_data.get('status')
        user = filter_form.cleaned_data.get('user')
        date_from = filter_form.cleaned_data.get('date_from')
        date_to = filter_form.cleaned_data.get('date_to')
        
        if role:
            shifts = shifts.filter(user__role=role)
        if status:
            shifts = shifts.filter(status=status)
        if user:
            shifts = shifts.filter(user=user)
        if date_from:
            shifts = shifts.filter(date__gte=date_from)
        if date_to:
            shifts = shifts.filter(date__lte=date_to)
    else:
        # Default: show next 14 days
        today = timezone.now().date()
        end_date = today + timedelta(days=13)
        shifts = shifts.filter(date__range=[today, end_date])
    
    shifts = shifts.order_by('date', 'start_time', 'user__role')
    
    # Organize shifts by date and user for calendar view
    shifts_by_date = defaultdict(lambda: defaultdict(list))
    dates = []
    
    for shift in shifts:
        if shift.date not in dates:
            dates.append(shift.date)
        shifts_by_date[shift.date][shift.user.id].append(shift)
    
        # Get all staff for the calendar columns
    staff_members = User.objects.filter(
        is_active=True,
        role__in=['doctor', 'therapist', 'nurse']
    ).order_by('role', 'last_name', 'first_name')
    
    # Count staff by role
    doctors_count = staff_members.filter(role='doctor').count()
    therapists_count = staff_members.filter(role='therapist').count()
    nurses_count = staff_members.filter(role='nurse').count()
    
    context = {
        'filter_form': filter_form,
        'shifts': shifts,
        'shifts_by_date': dict(shifts_by_date),
        'dates': sorted(dates),
        'staff_members': staff_members,
        'total_shifts': shifts.count(),
        'doctors_count': doctors_count,
        'therapists_count': therapists_count,
        'nurses_count': nurses_count,
    }
    
    return render(request, 'dashboard/shift_management.html', context)


@login_required
def shift_create(request):
    """Create a new shift"""
    if request.user.role != 'admin':
        raise PermissionDenied("您沒有權限訪問此頁面")
    
    if request.method == 'POST':
        form = ShiftForm(request.POST)
        if form.is_valid():
            shift = form.save()
            
            # Log the action
            AuditLog.objects.create(
                user=request.user,
                action='create',
                resource_type='Shift',
                resource_id=str(shift.id),
                ip_address=request.META.get('REMOTE_ADDR'),
                details=f'建立班表: {shift.user.get_full_name()} - {shift.date}'
            )
            
            messages.success(request, f'成功建立班表：{shift.user.get_full_name()} - {shift.date}')
            return redirect('dashboard:shift_management')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        form = ShiftForm()
    
    return render(request, 'dashboard/shift_form.html', {'form': form, 'action': 'create'})


@login_required
def shift_edit(request, shift_id):
    """Edit an existing shift"""
    if request.user.role != 'admin':
        raise PermissionDenied("您沒有權限訪問此頁面")
    
    shift = get_object_or_404(Shift, id=shift_id)
    
    if request.method == 'POST':
        form = ShiftForm(request.POST, instance=shift)
        if form.is_valid():
            shift = form.save()
            
            # Log the action
            AuditLog.objects.create(
                user=request.user,
                action='update',
                resource_type='Shift',
                resource_id=str(shift.id),
                ip_address=request.META.get('REMOTE_ADDR'),
                details=f'更新班表: {shift.user.get_full_name()} - {shift.date}'
            )
            
            messages.success(request, '成功更新班表')
            return redirect('dashboard:shift_management')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        form = ShiftForm(instance=shift)
    
    return render(request, 'dashboard/shift_form.html', {'form': form, 'action': 'edit', 'shift': shift})


@login_required
def shift_delete(request, shift_id):
    """Delete a shift"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': '沒有權限'}, status=403)
    
    shift = get_object_or_404(Shift, id=shift_id)
    
    if request.method == 'POST':
        shift_info = f'{shift.user.get_full_name()} - {shift.date}'
        shift.delete()
        
        # Log the action
        AuditLog.objects.create(
            user=request.user,
            action='delete',
            resource_type='Shift',
            resource_id=str(shift_id),
            ip_address=request.META.get('REMOTE_ADDR'),
            details=f'刪除班表: {shift_info}'
        )
        
        return JsonResponse({'success': True, 'message': '成功刪除班表'})
    
    return JsonResponse({'success': False, 'error': '無效的請求'}, status=400)


@login_required
def shift_bulk_action(request):
    """Handle bulk actions on shifts"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': '沒有權限'}, status=403)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            action = data.get('action')
            shift_ids = data.get('shift_ids', [])
            
            if not shift_ids:
                return JsonResponse({'success': False, 'error': '未選擇任何班表'})
            
            shifts = Shift.objects.filter(id__in=shift_ids)
            count = shifts.count()
            
            if action == 'confirm':
                shifts.update(status='confirmed')
                message = f'成功確認 {count} 個班表'
            elif action == 'cancel':
                shifts.update(status='cancelled')
                message = f'成功取消 {count} 個班表'
            elif action == 'delete':
                shifts.delete()
                message = f'成功刪除 {count} 個班表'
            else:
                return JsonResponse({'success': False, 'error': '無效的操作'})
            
            # Log the action
            AuditLog.objects.create(
                user=request.user,
                action='update',
                resource_type='Shift',
                ip_address=request.META.get('REMOTE_ADDR'),
                details=f'批次操作: {action} - {count} 個班表'
            )
            
            return JsonResponse({'success': True, 'message': message})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': '無效的請求'}, status=400)


@login_required
def user_management(request):
    """User management page for admin"""
    if request.user.role != 'admin':
        raise PermissionDenied("您沒有權限訪問此頁面")
    
    # Get filter parameters
    filter_form = UserFilterForm(request.GET or None)
    
    # Base queryset
    users = User.objects.all().order_by('-created_at')
    
    # Apply filters
    if filter_form.is_valid():
        role = filter_form.cleaned_data.get('role')
        is_active = filter_form.cleaned_data.get('is_active')
        search = filter_form.cleaned_data.get('search')
        
        if role:
            users = users.filter(role=role)
        if is_active:
            users = users.filter(is_active=(is_active == '1'))
        if search:
            users = users.filter(
                Q(username__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search)
            )
    
    # Get statistics
    total_users = User.objects.count()
    active_users = User.objects.filter(is_active=True).count()
    users_by_role = User.objects.values('role').annotate(count=Count('id'))
    
    context = {
        'filter_form': filter_form,
        'users': users,
        'total_users': total_users,
        'active_users': active_users,
        'users_by_role': {item['role']: item['count'] for item in users_by_role},
    }
    
    return render(request, 'dashboard/user_management.html', context)


@login_required
def user_create(request):
    """Create a new user"""
    if request.user.role != 'admin':
        raise PermissionDenied("您沒有權限訪問此頁面")
    
    if request.method == 'POST':
        form = UserCreateForm(request.POST)
        if form.is_valid():
            user = form.save()
            
            # Log the action
            AuditLog.objects.create(
                user=request.user,
                action='create',
                resource_type='User',
                resource_id=str(user.id),
                ip_address=request.META.get('REMOTE_ADDR'),
                details=f'建立使用者: {user.username} ({user.get_role_display()})'
            )
            
            messages.success(request, f'成功建立使用者：{user.username}')
            return redirect('dashboard:user_management')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = UserCreateForm()
    
    return render(request, 'dashboard/user_form.html', {'form': form, 'action': 'create'})


@login_required
def user_edit(request, user_id):
    """Edit an existing user"""
    if request.user.role != 'admin':
        raise PermissionDenied("您沒有權限訪問此頁面")
    
    user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        form = UserManagementForm(request.POST, instance=user)
        if form.is_valid():
            user = form.save()
            
            # Log the action
            AuditLog.objects.create(
                user=request.user,
                action='update',
                resource_type='User',
                resource_id=str(user.id),
                ip_address=request.META.get('REMOTE_ADDR'),
                details=f'更新使用者: {user.username}'
            )
            
            messages.success(request, '成功更新使用者資料')
            return redirect('dashboard:user_management')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = UserManagementForm(instance=user)
    
    return render(request, 'dashboard/user_form.html', {
        'form': form, 
        'action': 'edit', 
        'edit_user': user
    })


@login_required
def user_delete(request, user_id):
    """Delete a user"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': '沒有權限'}, status=403)
    
    user = get_object_or_404(User, id=user_id)
    
    # Prevent deleting yourself
    if user.id == request.user.id:
        return JsonResponse({'success': False, 'error': '無法刪除自己的帳號'}, status=400)
    
    if request.method == 'POST':
        username = user.username
        user.delete()
        
        # Log the action
        AuditLog.objects.create(
            user=request.user,
            action='delete',
            resource_type='User',
            resource_id=str(user_id),
            ip_address=request.META.get('REMOTE_ADDR'),
            details=f'刪除使用者: {username}'
        )
        
        return JsonResponse({'success': True, 'message': '成功刪除使用者'})
    
    return JsonResponse({'success': False, 'error': '無效的請求'}, status=400)


@login_required
def user_toggle_active(request, user_id):
    """Toggle user active status"""
    if request.user.role != 'admin':
        return JsonResponse({'success': False, 'error': '沒有權限'}, status=403)
    
    user = get_object_or_404(User, id=user_id)
    
    # Prevent deactivating yourself
    if user.id == request.user.id:
        return JsonResponse({'success': False, 'error': '無法停用自己的帳號'}, status=400)
    
    if request.method == 'POST':
        user.is_active = not user.is_active
        user.save()
        
        status_text = '啟用' if user.is_active else '停用'
        
        # Log the action
        AuditLog.objects.create(
            user=request.user,
            action='update',
            resource_type='User',
            resource_id=str(user.id),
            ip_address=request.META.get('REMOTE_ADDR'),
            details=f'{status_text}使用者: {user.username}'
        )
        
        return JsonResponse({
            'success': True, 
            'message': f'已{status_text}使用者',
            'is_active': user.is_active
        })
    
    return JsonResponse({'success': False, 'error': '無效的請求'}, status=400)


@login_required
def user_reset_password(request, user_id):
    """Reset user password"""
    if request.user.role != 'admin':
        raise PermissionDenied("您沒有權限訪問此頁面")
    
    user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        form = PasswordResetFormAdmin(request.POST)
        if form.is_valid():
            new_password = form.cleaned_data['new_password1']
            user.set_password(new_password)
            user.save()
            
            # Log the action
            AuditLog.objects.create(
                user=request.user,
                action='update',
                resource_type='User',
                resource_id=str(user.id),
                ip_address=request.META.get('REMOTE_ADDR'),
                details=f'重設使用者密碼: {user.username}'
            )
            
            messages.success(request, f'已成功重設 {user.username} 的密碼')
            return redirect('dashboard:user_management')
        else:
            for error in form.non_field_errors():
                messages.error(request, error)
    else:
        form = PasswordResetFormAdmin()
    
    return render(request, 'dashboard/user_reset_password.html', {
        'form': form,
        'reset_user': user
    })


@login_required
def shift_upload_excel(request):
    """Upload Excel file to bulk create shifts"""
    if request.user.role != 'admin':
        raise PermissionDenied("您沒有權限訪問此頁面")
    
    if request.method == 'POST':
        form = ShiftExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            try:
                # Load workbook
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                
                # Track results
                created_count = 0
                error_count = 0
                errors = []
                
                # Process rows (skip header row)
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        # Expected columns: 員工帳號, 班別, 日期, 開始時間, 結束時間, 地點, 備註
                        username = row[0]
                        shift_type = row[1]
                        date_val = row[2]
                        start_time_val = row[3]
                        end_time_val = row[4]
                        location = row[5] if len(row) > 5 else ''
                        notes = row[6] if len(row) > 6 else ''
                        
                        # Skip empty rows
                        if not username or not date_val:
                            continue
                        
                        # Find user
                        try:
                            user = User.objects.get(username=username, is_active=True)
                        except User.DoesNotExist:
                            errors.append(f'第 {row_num} 行：找不到使用者 "{username}"')
                            error_count += 1
                            continue
                        
                        # Parse date
                        if isinstance(date_val, datetime):
                            shift_date = date_val.date()
                        elif isinstance(date_val, str):
                            try:
                                shift_date = datetime.strptime(date_val, '%Y-%m-%d').date()
                            except ValueError:
                                try:
                                    shift_date = datetime.strptime(date_val, '%Y/%m/%d').date()
                                except ValueError:
                                    errors.append(f'第 {row_num} 行：日期格式錯誤 "{date_val}"')
                                    error_count += 1
                                    continue
                        else:
                            errors.append(f'第 {row_num} 行：無效的日期格式')
                            error_count += 1
                            continue
                        
                        # Parse start time
                        if isinstance(start_time_val, datetime):
                            start_time = start_time_val.time()
                        elif isinstance(start_time_val, dt_time):
                            start_time = start_time_val
                        elif isinstance(start_time_val, str):
                            try:
                                start_time = datetime.strptime(start_time_val, '%H:%M').time()
                            except ValueError:
                                errors.append(f'第 {row_num} 行：開始時間格式錯誤 "{start_time_val}"')
                                error_count += 1
                                continue
                        else:
                            errors.append(f'第 {row_num} 行：無效的開始時間格式')
                            error_count += 1
                            continue
                        
                        # Parse end time
                        if isinstance(end_time_val, datetime):
                            end_time = end_time_val.time()
                        elif isinstance(end_time_val, dt_time):
                            end_time = end_time_val
                        elif isinstance(end_time_val, str):
                            try:
                                end_time = datetime.strptime(end_time_val, '%H:%M').time()
                            except ValueError:
                                errors.append(f'第 {row_num} 行：結束時間格式錯誤 "{end_time_val}"')
                                error_count += 1
                                continue
                        else:
                            errors.append(f'第 {row_num} 行：無效的結束時間格式')
                            error_count += 1
                            continue
                        
                        # Validate shift type
                        valid_shift_types = dict(Shift.SHIFT_TYPE_CHOICES).keys()
                        if shift_type not in valid_shift_types:
                            errors.append(f'第 {row_num} 行：無效的班別 "{shift_type}"，有效值為：{", ".join(valid_shift_types)}')
                            error_count += 1
                            continue
                        
                        # Check for overlapping shifts
                        overlapping = Shift.objects.filter(
                            user=user,
                            date=shift_date,
                            status__in=['scheduled', 'confirmed']
                        ).exclude(
                            end_time__lte=start_time
                        ).exclude(
                            start_time__gte=end_time
                        )
                        
                        if overlapping.exists():
                            errors.append(f'第 {row_num} 行：班表時間衝突 ({user.get_full_name()} - {shift_date})')
                            error_count += 1
                            continue
                        
                        # Create shift
                        Shift.objects.create(
                            user=user,
                            shift_type=shift_type,
                            date=shift_date,
                            start_time=start_time,
                            end_time=end_time,
                            location=location or '',
                            notes=notes or '',
                            status='scheduled'
                        )
                        created_count += 1
                        
                    except Exception as e:
                        errors.append(f'第 {row_num} 行：處理錯誤 - {str(e)}')
                        error_count += 1
                        continue
                
                # Log the action
                AuditLog.objects.create(
                    user=request.user,
                    action='bulk_create',
                    resource_type='Shift',
                    resource_id='excel_upload',
                    ip_address=request.META.get('REMOTE_ADDR'),
                    details=f'Excel 批次匯入班表：成功 {created_count} 筆，失敗 {error_count} 筆'
                )
                
                # Show results
                if created_count > 0:
                    messages.success(request, f'成功匯入 {created_count} 筆班表')
                
                if error_count > 0:
                    error_summary = '<br>'.join(errors[:10])  # Show first 10 errors
                    if len(errors) > 10:
                        error_summary += f'<br>... 以及其他 {len(errors) - 10} 個錯誤'
                    messages.warning(request, f'有 {error_count} 筆資料匯入失敗：<br>{error_summary}')
                
                if created_count > 0:
                    return redirect('dashboard:shift_management')
                
            except Exception as e:
                messages.error(request, f'讀取 Excel 檔案時發生錯誤：{str(e)}')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        form = ShiftExcelUploadForm()
    
    return render(request, 'dashboard/shift_upload_excel.html', {
        'form': form,
        'title': '匯入班表 Excel'
    })


@login_required
def shift_download_template(request):
    """Download Excel template for shift upload"""
    if request.user.role != 'admin':
        raise PermissionDenied("您沒有權限訪問此頁面")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "班表範本"
    
    # Add headers
    headers = ['員工帳號', '班別', '日期', '開始時間', '結束時間', '地點', '備註']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
    
    # Add example data with multiple shift types
    example_data = [
        ['doctor1', 'morning', '2026-02-01', '08:00', '12:00', '門診一', '一般門診'],
        ['doctor1', 'afternoon', '2026-02-01', '14:00', '17:00', '門診一', '特診'],
        ['therapist1', 'morning', '2026-02-01', '09:00', '12:00', '復健科', '機器人治療'],
        ['therapist1', 'afternoon', '2026-02-01', '13:00', '17:00', '復健科', ''],
        ['nurse1', 'morning', '2026-02-01', '08:00', '16:00', '護理站', '日班'],
        ['nurse1', 'night', '2026-02-02', '00:00', '08:00', '護理站', '夜班'],
    ]
    
    for row_data in example_data:
        ws.append(row_data)
    
    # Add instructions sheet
    ws2 = wb.create_sheet("填寫說明")
    instructions = [
        ['欄位名稱', '說明', '格式範例', '必填'],
        ['員工帳號', '系統中的使用者帳號（username）', 'doctor1, therapist1', '是'],
        ['班別', '班別類型', 'morning, afternoon, evening, night, on_call', '是'],
        ['日期', '班表日期', '2026-02-01 或 2026/02/01', '是'],
        ['開始時間', '上班時間', '08:00', '是'],
        ['結束時間', '下班時間', '17:00', '是'],
        ['地點', '工作地點', '門診一、復健科', '否'],
        ['備註', '其他備註', '特殊說明', '否'],
        [],
        ['班別代碼說明：'],
        ['morning', '早班'],
        ['afternoon', '午班'],
        ['evening', '晚班'],
        ['night', '夜班'],
        ['on_call', '待命'],
    ]
    
    for row_data in instructions:
        ws2.append(row_data)
    
    # Style instruction headers
    for cell in ws2[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    
    # Adjust column widths
    for ws_temp in [ws, ws2]:
        for column in ws_temp.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_temp.column_dimensions[column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="shift_schedule_template.xlsx"'
    
    wb.save(response)
    
    # Log the action
    AuditLog.objects.create(
        user=request.user,
        action='download',
        resource_type='Template',
        resource_id='shift_excel_template',
        ip_address=request.META.get('REMOTE_ADDR'),
        details='下載班表 Excel 範本'
    )
    
    return response